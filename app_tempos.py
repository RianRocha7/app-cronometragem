import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import time
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage, ImageOps

st.set_page_config(page_title="App de Tempos", layout="wide")

st.markdown("""
<style>
@media (max-width: 768px) {
    /* Força os blocos de 6 colunas (nossa matriz) a virarem um carrossel horizontal no celular */
    [data-testid="stHorizontalBlock"]:has(> div:nth-child(6)) {
        flex-wrap: nowrap !important;
        overflow-x: auto !important;
        -webkit-overflow-scrolling: touch;
        padding-bottom: 5px;
    }
    [data-testid="stHorizontalBlock"]:has(> div:nth-child(6)) > div {
        min-width: 55px !important;
    }
    [data-testid="stHorizontalBlock"]:has(> div:nth-child(6)) button {
        padding: 0.1rem !important;
        font-size: 0.75rem !important;
        min-height: 2rem !important;
    }
    [data-testid="stHorizontalBlock"]:has(> div:nth-child(6)) p {
        font-size: 0.75rem !important;
        text-align: center;
        margin-bottom: 0px !important;
    }
}
</style>
""", unsafe_allow_html=True)

if 'df_tempos' not in st.session_state:
    st.session_state.df_tempos = pd.DataFrame(columns=["Nº", "Tempo (min cent)", "Pares", "Incluir?"])
if 'inicio' not in st.session_state:
    st.session_state.inicio = None
if 'tempo_acumulado' not in st.session_state:
    st.session_state.tempo_acumulado = 0.0
if 'status' not in st.session_state:
    st.session_state.status = 'parado'
if 'roteiro' not in st.session_state:
    st.session_state.roteiro = []
if 'esforco_idx' not in st.session_state:
    st.session_state.esforco_idx = 2 
if 'habilidade_idx' not in st.session_state:
    st.session_state.habilidade_idx = 2

with st.sidebar:
    st.header("📦 Envelope do Modelo")
    st.markdown("Construa o roteiro de operações aqui.")
    
    ref_modelo = st.text_input("Referência / Nome do Modelo")
    foto_modelo = st.file_uploader("Buscar Foto do Modelo", type=['png', 'jpg', 'jpeg', 'heic'])
    
    if foto_modelo:
        st.image(foto_modelo, caption=ref_modelo, use_container_width=True)
        
    st.divider()
    st.subheader("Operações Salvas")
    
    if st.session_state.roteiro:
        df_roteiro = pd.DataFrame(st.session_state.roteiro)
        
        df_exibicao = df_roteiro[["Operação", "TP"]].copy()
        df_exibicao["TP"] = df_exibicao["TP"].apply(lambda x: f"{x:.3f}".replace('.', ','))
        st.dataframe(df_exibicao, hide_index=True, use_container_width=True)
        
        st.markdown("### Exportar Roteiro")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Roteiro"
        
        ws["A1"] = "DATA:"
        ws["B1"] = datetime.now().strftime("%d/%m/%Y")
        ws["A2"] = "MODELO:"
        ws["B2"] = ref_modelo
        ws["A3"] = "SETOR:"
        ws["B3"] = st.session_state.roteiro[0]["Setor"] if st.session_state.roteiro else ""
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        
        if foto_modelo:
            try:
                # Tratamento avançado de imagem para fotos de celular
                img_pil = PILImage.open(foto_modelo)
                img_pil = ImageOps.exif_transpose(img_pil) # Corrige a rotação do celular
                img_pil = img_pil.convert('RGB') # Remove fundo transparente ou incompatibilidade do iOS/Android
                img_pil.thumbnail((250, 250)) 
                
                img_byte_arr = io.BytesIO()
                img_pil.save(img_byte_arr, format='PNG')
                img_excel = OpenpyxlImage(io.BytesIO(img_byte_arr.getvalue()))
                ws.add_image(img_excel, "D1")
            except Exception as e:
                st.error(f"Erro ao processar foto: {str(e)}")
        
        linha_atual = 5
        ws[f"A{linha_atual}"] = "DESCRIÇÃO DO PROCESSO"
        ws[f"B{linha_atual}"] = "TEMPO PADRÃO (min)"
        linha_atual += 1
        
        for op in st.session_state.roteiro:
            ws[f"A{linha_atual}"] = op["Operação"]
            ws[f"B{linha_atual}"] = round(op["TP"], 3)
            linha_atual += 1
            
        buffer = io.BytesIO()
        wb.save(buffer)
        
        st.download_button(
            label="📥 Baixar Roteiro em Excel",
            data=buffer.getvalue(),
            file_name=f"Roteiro_{ref_modelo if ref_modelo else 'Modelo'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        
        if st.button("🗑️ Limpar Envelope", use_container_width=True):
            st.session_state.roteiro = []
            st.rerun()
    else:
        st.info("Nenhuma operação salva no envelope ainda.")

st.title("⏱️ Folha de Tempos Digital")
st.divider()

col1, col2 = st.columns([1, 1.4])

with col1:
    st.subheader("Dados da Operação")
    operacao = st.text_input("Descrição da Operação", placeholder="Ex: Costura de Gáspea")
    
    # Setores atualizados, Costura é o padrão (índice 1)
    lista_setores = ["Corte", "Costura", "Pré-Fabricado", "Montagem", "Acabamento", "MANUAL", "AMOSTRAS"]
    setor = st.selectbox("Setor", lista_setores, index=1) 
    
    usa_maquina = st.checkbox("⚙️ Operação utiliza máquina?", value=True, help="Marcado = 12% Quebra. Desmarcado = 10% Quebra.")

with col2:
    st.subheader("Avaliação (Habilidade e Esforço)")
    st.markdown("<span style='font-size: 0.9em; color: gray;'>Selecione a porcentagem correspondente:</span>", unsafe_allow_html=True)
    
    labels_hab = ["SUPER", "EXCELENTE", "MÉDIO", "REGULAR", "POBRE"]
    labels_esf = ["SUPER", "EXCELENTE", "MÉDIO", "REGULAR", "POBRE"]
    
    matriz_valores = [
        [1.30, 1.22, 1.13, 1.02, 0.88], 
        [1.23, 1.16, 1.07, 0.96, 0.83], 
        [1.15, 1.08, 1.00, 0.90, 0.78], 
        [1.06, 0.98, 0.92, 0.83, 0.72], 
        [0.95, 0.90, 0.83, 0.75, 0.65]  
    ]
    
    cols_cabecalho = st.columns([1.2, 1, 1, 1, 1, 1])
    cols_cabecalho[0].markdown("**ESF \ HAB**")
    for j, hab in enumerate(labels_hab):
        cols_cabecalho[j+1].markdown(f"**<div style='text-align: center;'>{hab}</div>**", unsafe_allow_html=True)
        
    for i, esf in enumerate(labels_esf):
        cols_linha = st.columns([1.2, 1, 1, 1, 1, 1])
        cols_linha[0].markdown(f"<div style='margin-top: 10px;'><b>{esf}</b></div>", unsafe_allow_html=True)
        for j, hab in enumerate(labels_hab):
            valor_pct = int(matriz_valores[i][j] * 100)
            is_selected = (st.session_state.esforco_idx == i and st.session_state.habilidade_idx == j)
            tipo_btn = "primary" if is_selected else "secondary"
            
            if cols_linha[j+1].button(f"{valor_pct}%", key=f"btn_{i}_{j}", type=tipo_btn, use_container_width=True):
                st.session_state.esforco_idx = i
                st.session_state.habilidade_idx = j
                st.rerun()

    eficiencia = matriz_valores[st.session_state.esforco_idx][st.session_state.habilidade_idx]

st.divider()

st.subheader("Controle do Cronômetro")

html_cronometro = f"""
<div style="text-align: center; font-family: 'Courier New', Courier, monospace; border: 2px solid #ccc; border-radius: 10px; padding: 20px; margin-bottom: 20px; background-color: rgba(128, 128, 128, 0.1);">
    <div id="display_sec" style="font-size: 3.5rem; font-weight: bold; line-height: 1;">00:00.00</div>
    <div id="display_cent" style="font-size: 1.5rem; color: #888; margin-top: 5px;">0.000 min cent</div>
</div>
<script>
    var status = '{st.session_state.status}';
    var accumulated = {st.session_state.tempo_acumulado};
    var python_start = {st.session_state.inicio if st.session_state.inicio else 0};
    
    function updateTimer() {{
        var now = Date.now() / 1000;
        var current = accumulated;
        if (status === 'rodando') {{
            current += (now - python_start);
        }}
        
        var total_ms = Math.floor(current * 1000);
        var mins = Math.floor(total_ms / 60000);
        var secs = Math.floor((total_ms % 60000) / 1000);
        var ms = Math.floor((total_ms % 1000) / 10);
        
        var display_sec = (mins < 10 ? "0" + mins : mins) + ":" + 
                          (secs < 10 ? "0" + secs : secs) + "." + 
                          (ms < 10 ? "0" + ms : ms);
        document.getElementById('display_sec').innerText = display_sec;
        document.getElementById('display_cent').innerText = (current / 60).toFixed(3) + " min cent";
    }}
    updateTimer();
    if (status === 'rodando') {{ setInterval(updateTimer, 40); }}
</script>
"""
components.html(html_cronometro, height=160)

col_btn1, col_btn2, col_btn3 = st.columns(3)

with col_btn1:
    if st.session_state.status == 'rodando':
        if st.button("⏸️ Pausar Relógio", use_container_width=True):
            agora = time.time()
            st.session_state.tempo_acumulado += (agora - st.session_state.inicio)
            st.session_state.status = 'pausado'
            st.rerun()
    else:
        texto_iniciar = "▶️ Continuar" if st.session_state.status == 'pausado' else "▶️ Iniciar"
        if st.button(texto_iniciar, use_container_width=True):
            st.session_state.inicio = time.time()
            st.session_state.status = 'rodando'
            st.rerun()

with col_btn2:
    if st.button("⏱️ Registrar Tomada", use_container_width=True):
        if st.session_state.status == 'rodando':
            agora = time.time()
            tempo_total = st.session_state.tempo_acumulado + (agora - st.session_state.inicio)
            
            novo_dado = {
                "Nº": len(st.session_state.df_tempos) + 1, 
                "Tempo (min cent)": tempo_total / 60.0, 
                "Pares": 1.0, 
                "Incluir?": True
            }
            st.session_state.df_tempos = pd.concat([st.session_state.df_tempos, pd.DataFrame([novo_dado])], ignore_index=True)
            st.session_state.tempo_acumulado = 0.0
            st.session_state.inicio = time.time()
            st.rerun()
            
        elif st.session_state.status == 'pausado':
            tempo_total = st.session_state.tempo_acumulado
            if tempo_total > 0:
                novo_dado = {
                    "Nº": len(st.session_state.df_tempos) + 1, 
                    "Tempo (min cent)": tempo_total / 60.0, 
                    "Pares": 1.0, 
                    "Incluir?": True
                }
                st.session_state.df_tempos = pd.concat([st.session_state.df_tempos, pd.DataFrame([novo_dado])], ignore_index=True)
                st.session_state.tempo_acumulado = 0.0
                st.rerun()
            else:
                st.warning("O cronômetro está zerado.")
        else:
            st.warning("Inicie o cronômetro antes de registrar a tomada.")

with col_btn3:
    if st.button("🔄 Zerar Relógio", use_container_width=True):
        st.session_state.df_tempos = pd.DataFrame(columns=["Nº", "Tempo (min cent)", "Pares", "Incluir?"])
        st.session_state.tempo_acumulado = 0.0
        st.session_state.status = 'parado'
        st.rerun()

if not st.session_state.df_tempos.empty:
    st.divider()
    st.subheader("Análise de Tempos da Operação")
    
    st.write("Ajuste a quantidade de 'Pares' individualmente na tabela, ou aplique a todas as tomadas de uma vez:")
    
    col_p1, col_p2, col_p3 = st.columns([1, 1, 2])
    with col_p1:
        pares_massa = st.number_input("Pares por tomada:", min_value=0.1, step=0.5, value=1.0, format="%.1f")
    with col_p2:
        st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True) # Alinhamento vertical
        if st.button("🔄 Aplicar a Todas"):
            st.session_state.df_tempos["Pares"] = pares_massa
            st.rerun()
            
    st.session_state.df_tempos = st.data_editor(
        st.session_state.df_tempos, 
        use_container_width=True, 
        hide_index=True,
        column_config={
            "Nº": st.column_config.NumberColumn(disabled=True, width="small"),
            "Tempo (min cent)": st.column_config.NumberColumn(format="%.3f", disabled=True),
            "Pares": st.column_config.NumberColumn(
                help="Quantidade de pares fabricados nesta tomada",
                min_value=0.1,
                step=0.5,
                format="%.1f"
            ),
            "Incluir?": st.column_config.CheckboxColumn(width="small")
        }
    )
    
    df_filtrado = st.session_state.df_tempos[st.session_state.df_tempos["Incluir?"] == True].copy()
    
    if not df_filtrado.empty:
        # CÁLCULO AJUSTADO: Soma total dos tempos dividida pela soma total dos pares
        soma_tempos = df_filtrado["Tempo (min cent)"].sum()
        soma_pares = df_filtrado["Pares"].sum()
        tempo_medio = soma_tempos / soma_pares if soma_pares > 0 else 0.0
        
        tolerancia = 0.12 if usa_maquina else 0.10
        
        tn = tempo_medio * eficiencia
        tp = tn * (1 + tolerancia)
        
        st.markdown(f"**Eficiência Selecionada:** `{int(eficiencia * 100)}%` &nbsp;&nbsp;|&nbsp;&nbsp; **Quebra Considerada:** `{int(tolerancia * 100)}%`")
        
        st.markdown("### Resultados da Operação Atual")
        c1, c2, c3 = st.columns(3)
        c1.metric("Tempo Médio (TM) por Par", f"{tempo_medio:.3f}".replace('.', ',') + " min")
        c2.metric("Tempo Normal (TN)", f"{tn:.3f}".replace('.', ',') + " min")
        c3.metric("Tempo Padrão (TP)", f"{tp:.3f}".replace('.', ',') + " min")
        
        st.divider()
        if st.button("💾 Adicionar ao Envelope (Roteiro)", use_container_width=True, type="primary"):
            if not operacao:
                st.error("Por favor, preencha a 'Descrição da Operação' antes de salvar.")
            else:
                nova_operacao = {
                    "Operação": operacao,
                    "Setor": setor,
                    "TP": tp
                }
                st.session_state.roteiro.append(nova_operacao)
                
                # Zera o cronômetro para iniciar a próxima operação na fábrica
                st.session_state.df_tempos = pd.DataFrame(columns=["Nº", "Tempo (min cent)", "Pares", "Incluir?"])
                st.session_state.tempo_acumulado = 0.0
                st.session_state.status = 'parado'
                st.success("Operação adicionada ao envelope com sucesso!")
                time.sleep(1.5)
                st.rerun()
    else:
        st.warning("Nenhuma tomada incluída no cálculo.")
